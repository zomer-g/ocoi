"""Importer for the Knesset "תקציב קשר עם הציבור" (constituent-outreach
expense) Excel reports.

The published files (`MKExpenses<year>.xlsx`) share a stable schema:

    Row 0:  banner — "הנתונים הינם גולמיים ואינם מבוקרים"
    Row 1:  blank
    Row 2:  header row — exact strings:
        שם חבר הכנסת | שם סעיף הוצאה | שם בית עסק/ ספק |
        תאריך ביצוע/ תאריך חשבונית | סכום בש"ח | פרטים/ הערות |
        אשראי | אסמכתאות לעסקה
    Row 3+: data rows. ~22-24K rows per file. Negative amounts are
            refunds. The 2024 file has 9 columns (last empty); the
            2025 file has 8 columns.

Some "suppliers" are actually internal Knesset services
("משיכה ממחסן הכנסת" וכו׳). The caller can map them to a single canonical
Company name via `KNESSET_INTERNAL_SUPPLIERS`.

This module is import-side only: it parses the file and yields typed rows.
Aggregation and DB persistence happen in
`ocoi_api.services.import_service.run_mk_expenses_import`.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Iterator

from openpyxl import load_workbook

from ocoi_common.models import MkExpenseRow

logger = logging.getLogger("ocoi.importer.mk_expenses")


# Header strings that identify the data sheet (compared with strip())
HEADER_MK_NAME = "שם חבר הכנסת"
HEADER_CATEGORY = "שם סעיף הוצאה"
HEADER_SUPPLIER = "שם בית עסק/ ספק"
HEADER_DATE = "תאריך ביצוע/ תאריך חשבונית"
HEADER_AMOUNT = 'סכום בש"ח'
HEADER_NOTES = "פרטים/ הערות"

REQUIRED_HEADER_FRAGMENTS = (HEADER_MK_NAME, HEADER_SUPPLIER, HEADER_AMOUNT)


# Internal Knesset "vendors" — these are not external suppliers, they're
# internal Knesset services / depots. The caller collapses all of them
# to one canonical Company "הכנסת".
KNESSET_INTERNAL_SUPPLIERS: frozenset[str] = frozenset({
    "משיכה ממחסן הכנסת",
    "שירותי דואר הכנסת",
    "שירותי הסעדה הכנסת",
    "מעבדת המחשבים",
    "כנסת ישראל",
    "מחסן הכנסת",
})

CANONICAL_KNESSET_NAME = "הכנסת"


def canonical_supplier_name(raw: str) -> str:
    """Map raw supplier strings to their canonical entity name. Internal
    Knesset services collapse to "הכנסת"; everything else passes through
    after a whitespace cleanup."""
    if not raw:
        return ""
    cleaned = " ".join(raw.split())  # collapse whitespace
    if cleaned in KNESSET_INTERNAL_SUPPLIERS:
        return CANONICAL_KNESSET_NAME
    return cleaned


def _normalise_mk_name(raw: str) -> str:
    if not raw:
        return ""
    return " ".join(raw.split())


def _coerce_amount(value) -> float:
    """The amount column is mostly numeric, occasionally a string. Anything
    we can't parse becomes 0.0 so the row still aggregates."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _coerce_date(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str) and value.strip():
        return value.strip()
    return None


def _looks_like_header(row: tuple) -> bool:
    """A row is the header row iff it contains the MK-name marker and the
    supplier marker side by side."""
    cells = [str(c).strip() if c is not None else "" for c in row]
    text = " | ".join(cells)
    return all(frag in text for frag in REQUIRED_HEADER_FRAGMENTS)


def _column_indices(header_row: tuple) -> dict[str, int]:
    """Map our logical field names to 0-based column indices using the
    header row text."""
    cells = [str(c).strip() if c is not None else "" for c in header_row]
    idx: dict[str, int] = {}
    for i, cell in enumerate(cells):
        if cell == HEADER_MK_NAME:
            idx["mk_name"] = i
        elif cell == HEADER_CATEGORY:
            idx["category"] = i
        elif cell == HEADER_SUPPLIER:
            idx["supplier"] = i
        elif cell == HEADER_DATE:
            idx["date"] = i
        elif cell == HEADER_AMOUNT:
            idx["amount"] = i
        elif cell == HEADER_NOTES:
            idx["notes"] = i
    return idx


def iter_rows(file_bytes: bytes) -> Iterator[MkExpenseRow]:
    """Stream `MkExpenseRow` records from one MK-expenses .xlsx file.

    Uses openpyxl in read_only mode so memory stays bounded — the data
    sheets have ~24K rows × 8 columns, well under any practical limit.
    Skips rows that don't have both an MK name and a supplier name.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row: tuple | None = None
            col_idx: dict[str, int] = {}

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                # Locate the header (typically row index 2 — but be lenient).
                if header_row is None:
                    if _looks_like_header(row):
                        header_row = row
                        col_idx = _column_indices(row)
                    continue

                # Skip blank rows and stop iterating sheets that lose shape
                if all(c is None or str(c).strip() == "" for c in row):
                    continue

                def _at(key: str):
                    i = col_idx.get(key)
                    return row[i] if i is not None and i < len(row) else None

                mk_name = _normalise_mk_name(str(_at("mk_name") or "").strip())
                supplier_raw = str(_at("supplier") or "").strip()
                if not mk_name or not supplier_raw:
                    continue

                yield MkExpenseRow(
                    mk_name=mk_name,
                    expense_category=(str(_at("category") or "").strip() or None),
                    raw_supplier_name=supplier_raw,
                    date=_coerce_date(_at("date")),
                    amount=_coerce_amount(_at("amount")),
                    notes=(str(_at("notes") or "").strip() or None),
                    row_idx=row_idx,
                    sheet_name=sheet_name,
                )

            # Only the first sheet that matches the header layout is the
            # data sheet; later sheets (notes / explanation) are ignored.
            if header_row is not None:
                break
    finally:
        try:
            wb.close()
        except Exception:
            pass
